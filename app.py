import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import datetime
import io
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# 로컬 모듈 임포트
from data_loader import get_stock_list, download_prices_chunked
from screener import run_screener, check_vcp_pattern, check_trend_template, calculate_returns, calculate_rs_ratings

# 가격 통화 포맷팅 함수
def fmt_curr(val, ticker):
    if pd.isna(val) or isinstance(val, str):
        return val
    if ticker.endswith('.KS') or ticker.endswith('.KQ'):
        return f"{val:,.0f}원"
    else:
        return f"${val:,.2f}"

# 페이지 설정
st.set_page_config(
    page_title="Mark Minervini Trend Template & VCP Screener",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 커스텀 CSS 스타일 정의 (어두운 테마 가독성 극대화 및 스타일링)
st.markdown("""
<style>
    .main-title {
        font-size: 2.1rem;
        font-weight: 700;
        color: #8AB4F8; /* 밝은 청색 계열 */
        margin-bottom: 0.2rem;
    }
    .sub-title {
        font-size: 1.05rem;
        color: #BDC1C6;
        margin-bottom: 1.8rem;
    }
    .metric-card {
        background-color: #202124;
        color: #F1F3F4;
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #8AB4F8;
        margin-bottom: 10px;
    }
    .status-pass {
        color: #0F9D58;
        font-weight: bold;
    }
    .status-fail {
        color: #D93025;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

# 헤더 영역
st.markdown('<div class="main-title">Mark Minervini Trend Template & VCP 스크리너</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">초성장주 발굴을 위한 마크 미너비니의 8대 상승 추세 템플릿(MTT) 조건 및 VCP(변동성 수축 패턴) 스크리너</div>', unsafe_allow_html=True)

# 기법 설명 익스팬더
with st.expander("ℹ️ 마크 미너비니(Mark Minervini) 투자 기법 및 조건 안내"):
    st.markdown("""
    **마크 미너비니(Mark Minervini)**는 주식시장 투자 챔피언십 우승자이자 책 *'초성장주 투자 공식(Trade Like a Stock Market Wizard)'*의 저자입니다.
    그는 강력한 상승 국면에 진입한 **상승 2단계(Stage 2 Uptrend)** 종목을 매수하여 짧은 시간 안에 폭발적인 수익을 내는 전략을 사용합니다.
    
    ### 📌 트렌드 템플릿(Trend Template) 8대 요건
    1. **현재 주가 > 150일 이동평균선** AND **현재 주가 > 200일 이동평균선**
    2. **150일 이동평균선 > 200일 이동평균선** (이평선 우상향 정배열 시작)
    3. **200일 이동평균선이 최소 1개월(22영업일) 동안 우상향 흐름** 유지
    4. **50일 이동평균선 > 150일 이동평균선** AND **50일 이동평균선 > 200일 이동평균선**
    5. **현재 주가 > 50일 이동평균선** (이평선 위에서 거래되는 강세 흐름)
    6. **현재 주가 >= 52주 신저가 대비 30% 이상** 높을 것 (바닥을 다지고 상승 추세로 돌아선 증거)
    7. **현재 주가 >= 52주 신고가 대비 25% 이내**에 위치 (전고점 돌파 대기 상태)
    8. **상대 강도(Relative Strength) Rating이 70 이상** (전체 시장 종목 중 수익률 상위 30% 이내)

    ### 🌀 변동성 수축 패턴(VCP, Volatility Contraction Pattern)
    상승 돌파 직전 매도 물량이 말라붙으면서 주가 변동폭(진폭)과 거래량이 점진적으로 줄어드는 현상입니다.
    본 프로그램은 **최근 60일(3달)** 동안 **3단계 진폭 수축(Amp3 > Amp2 > Amp1)** 여부를 확인하고, 
    마지막 조임 구간의 진폭이 설정치 이하로 수축하였는지, 거래량이 30일 평균 대비 감소했는지, 최근 20일 고점에 근접했는지를 검증합니다.
    """)

# 세션 상태(Session State) 변수 초기화
if 'screened_df' not in st.session_state:
    st.session_state.screened_df = None
if 'last_run_time' not in st.session_state:
    st.session_state.last_run_time = None
if 'market_type_used' not in st.session_state:
    st.session_state.market_type_used = None
if 'vcp_applied' not in st.session_state:
    st.session_state.vcp_applied = None
if 'rs_ratings' not in st.session_state:
    st.session_state.rs_ratings = {}

# 사이드바: 스크리닝 파라미터 구성
st.sidebar.header("⚙️ 스크리닝 조건 설정")

market_choice = st.sidebar.selectbox(
    "대상 시장 선택",
    ["코스피 (KOSPI)", "코스닥 (KOSDAQ)", "미국 S&P 500", "미국 NASDAQ"],
    index=0
)

rs_rating_thresh = st.sidebar.slider(
    "최소 상대 강도 (RS Rating)",
    min_value=50,
    max_value=99,
    value=70,
    step=1,
    help="전체 상장 종목 중 최근 1년 수익률 상위 백분위수 조건입니다. 미너비니 기본 요건은 70 이상(추천은 80~90 이상)입니다."
)

apply_vcp = st.sidebar.toggle(
    "변동성 수축 패턴(VCP) 필터 적용",
    value=True,
    help="체크하면 3단계 점진적 진폭 수축 및 거래량 감소, 돌파 임박 등의 VCP 요건을 필터링에 반영합니다."
)

with st.sidebar.expander("🌀 VCP 상세 조건 설정", expanded=apply_vcp):
    vcp_amp_limit = st.slider(
        "마지막 단계 최대 진폭 (Amp1)",
        min_value=3.0,
        max_value=20.0,
        value=10.0,
        step=0.5,
        format="%f%%",
        help="최근 10일간 최고가와 최저가 간의 최대 진폭 허용치입니다. 작을수록 단단히 밀착되어 수축한 상태를 의미합니다."
    ) / 100.0
    
    vol_dryup_ratio = st.slider(
        "거래량 메마름(Dry-up) 비율",
        min_value=50.0,
        max_value=100.0,
        value=80.0,
        step=5.0,
        format="%f%%",
        help="최근 5일 평균 거래량이 30일 평균 거래량 대비 몇 % 이하로 말라붙어야 하는지 설정합니다."
    ) / 100.0
    
    breakout_pct = st.slider(
        "돌파 임박 가격 비율",
        min_value=85.0,
        max_value=100.0,
        value=95.0,
        step=1.0,
        format="%f%%",
        help="현재 주가가 최근 20일간 최고가 대비 몇 % 영역 이상에 도달하여 전고점 돌파 직전이어야 하는지 결정합니다."
    ) / 100.0

with st.sidebar.expander("🛠️ 시스템 & 성능 설정", expanded=False):
    limit_tickers = st.number_input(
        "테스트용 종목 개수 제한 (0 = 제한 없음)",
        min_value=0,
        max_value=5000,
        value=0,
        step=50,
        help="빠른 테스트를 위해 대상 시장의 상위 N개 티커만 임포트하고 분석하려면 설정하세요."
    )
    
    chunk_size = st.number_input(
        "API 다운로드 청크 크기",
        min_value=10,
        max_value=500,
        value=150,
        step=50,
        help="yfinance API로 한 번에 배치 다운로드 요청을 보낼 종목 개수입니다. 안정적인 연결을 위해 150 전후를 추천합니다."
    )

start_screening = st.sidebar.button("🚀 스크리닝 시작", use_container_width=True)

# ----------------- 스크리닝 비즈니스 로직 구동 -----------------
if start_screening:
    market_map = {
        "코스피 (KOSPI)": "KS",
        "코스닥 (KOSDAQ)": "KQ",
        "미국 S&P 500": "SP",
        "미국 NASDAQ": "NQ"
    }
    market_code = market_map[market_choice]
    
    # 1. 상장 종목 목록 가져오기
    status_text = st.empty()
    status_text.info(f"⏳ {market_choice} 상장 종목 리스트를 불러오고 있습니다...")
    
    try:
        stock_df = get_stock_list(market_code)
    except Exception as e:
        st.error(f"종목 리스트 수집 중 에러가 발생했습니다: {e}")
        stock_df = pd.DataFrame()
        
    if not stock_df.empty:
        if limit_tickers > 0:
            status_text.warning(f"⚠️ 빠른 테스트를 위해 종목 개수를 상위 {limit_tickers}개로 제한합니다.")
            stock_df = stock_df.head(limit_tickers)
            
        tickers = stock_df['ticker'].tolist()
        
        # 2. 가격 데이터 배치 다운로드
        status_text.info(f"⏳ 총 {len(tickers)}개 종목의 과거 2년 가격 데이터를 다운로드 중입니다. (청크 크기: {chunk_size})...")
        progress_bar = st.progress(0.0)
        
        all_data = []
        total_chunks = (len(tickers) + chunk_size - 1) // chunk_size
        
        # data_loader.download_prices_chunked 내부 루프를 대시보드 프로그레스 바 연동을 위해 변형
        from data_loader import suppress_stderr
        
        success_download = True
        for i in range(0, len(tickers), chunk_size):
            chunk = tickers[i:i + chunk_size]
            chunk_num = i // chunk_size + 1
            status_text.info(f"⏳ 데이터 다운로드 중... [청크 {chunk_num}/{total_chunks}] (Rate limit 방지를 위해 대기 시간 포함)")
            progress_bar.progress(chunk_num / total_chunks)
            
            try:
                with suppress_stderr():
                    df_chunk = yf.download(
                        tickers=chunk, 
                        period="2y", 
                        interval="1d", 
                        group_by="ticker", 
                        auto_adjust=True, 
                        threads=True,
                        progress=False
                    )
                if not df_chunk.empty:
                    all_data.append(df_chunk)
                time.sleep(1.5)  # Rate limit 예방
            except Exception as e:
                time.sleep(2.0)
                continue
                
        progress_bar.empty()
        
        if all_data:
            price_data = pd.concat(all_data, axis=1)
            status_text.info("⏳ 이동평균선 및 상대 강도(RS) 계산 및 미너비니 규칙 필터링을 진행 중입니다...")
            
            try:
                # 상대강도 보존용 계산 수행 및 세션 저장
                returns_dict = calculate_returns(price_data, tickers)
                rs_ratings = calculate_rs_ratings(returns_dict)
                st.session_state.rs_ratings = rs_ratings
                
                screened_res = run_screener(
                    full_df=price_data,
                    stock_list_df=stock_df,
                    apply_vcp=apply_vcp,
                    rs_rating_thresh=rs_rating_thresh,
                    vcp_amp_limit=vcp_amp_limit,
                    vol_dryup_ratio=vol_dryup_ratio,
                    breakout_pct=breakout_pct
                )
                
                st.session_state.screened_df = screened_res
                st.session_state.last_run_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                st.session_state.market_type_used = market_choice
                st.session_state.vcp_applied = apply_vcp
                
                status_text.success("🎉 스크리닝 작업이 정상 완료되었습니다!")
            except Exception as e:
                status_text.error(f"스크리닝 필터 작업 중 오류 발생: {e}")
        else:
            status_text.error("종목의 역사적 가격 데이터 다운로드에 실패하여 분석을 수행하지 못했습니다.")
    else:
        status_text.error("대상 시장의 종목 리스트가 유효하지 않습니다.")

# ----------------- 결과 출력 및 탭 레이아웃 -----------------
tab1, tab2 = st.tabs(["🔍 스크리닝 결과", "🔬 개별 종목 분석기"])

with tab1:
    if st.session_state.screened_df is not None:
        st.success(f"📊 분석 결과 리포트 (실행: {st.session_state.last_run_time} | 대상: {st.session_state.market_type_used})")
        
        if st.session_state.screened_df.empty:
            st.warning("설정하신 조건을 충족하는 종목이 포착되지 않았습니다. 조건 범위(RS Rating 하한, Amp1 한계 등)를 넓혀 다시 시작해 보세요.")
        else:
            st.markdown(f"#### 🎯 포착된 종목 리스트 (총 {len(st.session_state.screened_df)}개)")
            
            # 대시보드 화면용 테이블 포맷팅 가공
            df_display = st.session_state.screened_df.copy()
            df_display['Current_Price'] = df_display.apply(lambda r: fmt_curr(r['Current_Price'], r['Ticker']), axis=1)
            df_display['MA_50'] = df_display.apply(lambda r: fmt_curr(r['MA_50'], r['Ticker']), axis=1)
            df_display['MA_150'] = df_display.apply(lambda r: fmt_curr(r['MA_150'], r['Ticker']), axis=1)
            df_display['MA_200'] = df_display.apply(lambda r: fmt_curr(r['MA_200'], r['Ticker']), axis=1)
            df_display['52W_High'] = df_display.apply(lambda r: fmt_curr(r['52W_High'], r['Ticker']), axis=1)
            df_display['52W_Low'] = df_display.apply(lambda r: fmt_curr(r['52W_Low'], r['Ticker']), axis=1)
            
            df_display['Pct_Below_High'] = df_display['Pct_Below_High'].map('{:.2f}%'.format)
            df_display['Pct_Above_Low'] = df_display['Pct_Above_Low'].map('{:.2f}%'.format)
            
            def format_vcp_col(val):
                if val == "N/A":
                    return val
                return f"{val:.2f}%"
            df_display['VCP_Amp1'] = df_display['VCP_Amp1'].apply(format_vcp_col)
            
            # 한글 컬럼 매핑 및 전시
            df_display_ko = df_display.rename(columns={
                'Ticker': '티커',
                'Name': '종목명',
                'Current_Price': '현재가',
                'RS_Rating': '상대강도(RS)',
                'VCP_Amp1': '마지막진폭(Amp1)',
                'Pct_Below_High': '신고가대비(%)',
                'Pct_Above_Low': '신저가대비(%)',
                'MA_50': '50일 이평',
                'MA_150': '150일 이평',
                'MA_200': '200일 이평',
                '52W_High': '52주 최고가',
                '52W_Low': '52주 최저가'
            })
            
            st.dataframe(df_display_ko, use_container_width=True)
            
            # --- 엑셀 저장 및 다운로드 구현 ---
            st.markdown("### 📥 엑셀 레포트 다운로드")
            
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_excel = st.session_state.screened_df.copy()
                df_excel_ko = df_excel.rename(columns={
                    'Ticker': '티커',
                    'Name': '종목명',
                    'Current_Price': '현재가',
                    'RS_Rating': '상대강도(RS)',
                    'VCP_Amp1': '마지막진폭(Amp1)',
                    'Pct_Below_High': '신고가대비(%)',
                    'Pct_Above_Low': '신저가대비(%)',
                    'MA_50': '50일 이평',
                    'MA_150': '150일 이평',
                    'MA_200': '200일 이평',
                    '52W_High': '52주 최고가',
                    '52W_Low': '52주 최저가'
                })
                
                df_excel_ko.to_excel(writer, index=False, sheet_name='MTT_VCP_Screened')
                worksheet = writer.sheets['MTT_VCP_Screened']
                
                max_col = worksheet.max_column
                max_row = worksheet.max_row
                
                # 헤더 서식 지정
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
                align_center = Alignment(horizontal="center", vertical="center")
                
                worksheet.row_dimensions[1].height = 28
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                
                # 데이터 서식 지정
                align_left = Alignment(horizontal="left", vertical="center")
                align_right = Alignment(horizontal="right", vertical="center")
                data_font = Font(name="맑은 고딕", size=10)
                
                for row_idx in range(2, max_row + 1):
                    worksheet.row_dimensions[row_idx].height = 20
                    ticker_val = str(worksheet.cell(row=row_idx, column=1).value or '')
                    is_kr = ticker_val.endswith('.KS') or ticker_val.endswith('.KQ')
                    
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = data_font
                        
                        col_name = df_excel_ko.columns[col_idx - 1]
                        val = cell.value
                        
                        if col_name in ['티커', '상대강도(RS)']:
                            cell.alignment = align_center
                        elif col_name == '종목명':
                            cell.alignment = align_left
                        elif col_name == '마지막진폭(Amp1)' and val == 'N/A':
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_right
                            if isinstance(val, (int, float)):
                                if col_name in ['신고가대비(%)', '신저가대비(%)', '마지막진폭(Amp1)']:
                                    cell.number_format = '0.00"%"'
                                elif col_name in ['현재가', '50일 이평', '150일 이평', '200일 이평', '52주 최고가', '52주 최저가']:
                                    cell.number_format = '#,##0' if is_kr else '#,##0.00'
                
                # 자동 열 너비 계산
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or '')
                        korean_chars = sum(1 for c in val_str if ord(c) > 128)
                        length = len(val_str) + korean_chars
                        if length > max_len:
                            max_len = length
                    worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 30)
                
                if max_row > 1:
                    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                    
            excel_data = excel_buffer.getvalue()
            
            market_code_map = {
                "코스피 (KOSPI)": "KS",
                "코스닥 (KOSDAQ)": "KQ",
                "미국 S&P 500": "SP",
                "미국 NASDAQ": "NQ"
            }
            market_suffix = market_code_map.get(st.session_state.market_type_used, "ALL")
            today_str = datetime.date.today().strftime('%Y-%m-%d')
            vcp_suffix = "-VCP" if st.session_state.vcp_applied else ""
            excel_filename = f"MarkMinerviniMTT-{market_suffix}{vcp_suffix}-{today_str}.xlsx"
            
            st.download_button(
                label="📥 스타일이 지정된 엑셀 보고서 다운로드 (.xlsx)",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            # --- 개별 종목 차트 시각화 구역 ---
            st.markdown("---")
            st.markdown("### 📊 포착 종목 변동성 수축(VCP) 및 추세 시각화 차트")
            
            selected_stock_name = st.selectbox(
                "차트로 분석할 종목을 포착 리스트에서 선택하세요:",
                options=st.session_state.screened_df['Name'].tolist()
            )
            
            if selected_stock_name:
                row = st.session_state.screened_df[st.session_state.screened_df['Name'] == selected_stock_name].iloc[0]
                selected_ticker = row['Ticker']
                
                st.markdown(f"#### 🔍 {selected_stock_name} ({selected_ticker}) 차트 상세 분석")
                
                # 주가 데이터 다운로드 (차트는 깔끔한 이평 조회를 위해 2년 정보 취득)
                with st.spinner(f"주가 이력 로드 중... ({selected_ticker})"):
                    df_chart = yf.download(selected_ticker, period="2y", progress=False)
                    if isinstance(df_chart.columns, pd.MultiIndex):
                        df_chart.columns = df_chart.columns.droplevel(1)
                    df_chart = df_chart.dropna(subset=['Close', 'High', 'Low', 'Volume'])
                
                if not df_chart.empty:
                    # 보조 지표 계산
                    df_chart['SMA_50'] = df_chart['Close'].rolling(window=50).mean()
                    df_chart['SMA_150'] = df_chart['Close'].rolling(window=150).mean()
                    df_chart['SMA_200'] = df_chart['Close'].rolling(window=200).mean()
                    df_chart['Vol_SMA_20'] = df_chart['Volume'].rolling(window=20).mean()
                    
                    df_chart['High_52w'] = df_chart['High'].rolling(window=252).max()
                    df_chart['Low_52w'] = df_chart['Low'].rolling(window=252).min()
                    
                    current_price = df_chart['Close'].iloc[-1]
                    ma50_val = df_chart['SMA_50'].iloc[-1]
                    ma150_val = df_chart['SMA_150'].iloc[-1]
                    ma200_val = df_chart['SMA_200'].iloc[-1]
                    ma200_prev_val = df_chart['SMA_200'].iloc[-22] if len(df_chart) >= 22 else df_chart['SMA_200'].iloc[0]
                    h52_val = df_chart['High_52w'].iloc[-1]
                    l52_val = df_chart['Low_52w'].iloc[-1]
                    
                    # 8대 규칙 수동 검토용 불리언 값 계산
                    cond1 = (current_price > ma150_val) and (current_price > ma200_val)
                    cond2 = ma150_val > ma200_val
                    cond3 = ma200_val > ma200_prev_val
                    cond4 = (ma50_val > ma150_val) and (ma50_val > ma200_val)
                    cond5 = current_price > ma50_val
                    cond6 = current_price >= (l52_val * 1.30)
                    cond7 = current_price >= (h52_val * 0.75)
                    stock_rs = st.session_state.rs_ratings.get(selected_ticker, 0)
                    cond8 = stock_rs >= rs_rating_thresh
                    
                    # 1. Plotly Subplot 구성
                    fig = make_subplots(
                        rows=2, cols=1,
                        shared_xaxes=True,
                        vertical_spacing=0.08,
                        row_heights=[0.7, 0.3]
                    )
                    
                    # 캔들스틱 차트 추가
                    fig.add_trace(
                        go.Candlestick(
                            x=df_chart.index,
                            open=df_chart['Open'],
                            high=df_chart['High'],
                            low=df_chart['Low'],
                            close=df_chart['Close'],
                            name="주가",
                            increasing_line_color='#EA4335',
                            decreasing_line_color='#4285F4'
                        ),
                        row=1, col=1
                    )
                    
                    # 이동평균선 3종 추가
                    fig.add_trace(
                        go.Scatter(x=df_chart.index, y=df_chart['SMA_50'], line=dict(color='#FBBC05', width=1.5), name="50일 MA"),
                        row=1, col=1
                    )
                    fig.add_trace(
                        go.Scatter(x=df_chart.index, y=df_chart['SMA_150'], line=dict(color='#34A853', width=1.5), name="150일 MA"),
                        row=1, col=1
                    )
                    fig.add_trace(
                        go.Scatter(x=df_chart.index, y=df_chart['SMA_200'], line=dict(color='#EA4335', width=2), name="200일 MA"),
                        row=1, col=1
                    )
                    
                    # 52주 최고가 및 최저가 수평 가이드 라인 추가 (점선)
                    fig.add_trace(
                        go.Scatter(x=df_chart.index, y=df_chart['High_52w'], line=dict(color='#BDC1C6', width=1, dash='dash'), name="52주 신고가선"),
                        row=1, col=1
                    )
                    fig.add_trace(
                        go.Scatter(x=df_chart.index, y=df_chart['Low_52w'], line=dict(color='#BDC1C6', width=1, dash='dot'), name="52주 신저가선"),
                        row=1, col=1
                    )
                    
                    # VCP 수축 구간 반투명 박스 및 텍스트 레이블 마킹 (최근 60일 기준 데이터 존재 시)
                    if len(df_chart) >= 60:
                        d_60 = df_chart.index[-60]
                        d_30 = df_chart.index[-30]
                        d_10 = df_chart.index[-10]
                        d_end = df_chart.index[-1]
                        
                        p3 = df_chart.iloc[-60:-30]
                        p2 = df_chart.iloc[-30:-10]
                        p1 = df_chart.iloc[-10:]
                        
                        amp3_calc = (p3['High'].max() - p3['Low'].min()) / p3['Low'].min() * 100.0
                        amp2_calc = (p2['High'].max() - p2['Low'].min()) / p2['Low'].min() * 100.0
                        amp1_calc = (p1['High'].max() - p1['Low'].min()) / p1['Low'].min() * 100.0
                        
                        # 60~30일 전 수축구간 3
                        fig.add_vrect(
                            x0=d_60, x1=d_30,
                            fillcolor="rgba(66, 133, 244, 0.05)",
                            line_width=0,
                            annotation_text=f"Amp3: {amp3_calc:.1f}%",
                            annotation_position="top left",
                            row=1, col=1
                        )
                        # 30~10일 전 수축구간 2
                        fig.add_vrect(
                            x0=d_30, x1=d_10,
                            fillcolor="rgba(251, 188, 5, 0.05)",
                            line_width=0,
                            annotation_text=f"Amp2: {amp2_calc:.1f}%",
                            annotation_position="top left",
                            row=1, col=1
                        )
                        # 최근 10일 조임구간 1
                        fig.add_vrect(
                            x0=d_10, x1=d_end,
                            fillcolor="rgba(52, 168, 83, 0.08)",
                            line_width=0,
                            annotation_text=f"Amp1: {amp1_calc:.1f}% (최종 수축)",
                            annotation_position="top left",
                            row=1, col=1
                        )
                    
                    # 2. 거래량 보조 차트 구성
                    # 상승일 거래량은 빨간색, 하락일 거래량은 파란색으로 표현
                    vol_colors = []
                    for idx in range(len(df_chart)):
                        if idx == 0:
                            vol_colors.append('#EA4335')
                        else:
                            if df_chart['Close'].iloc[idx] >= df_chart['Close'].iloc[idx-1]:
                                vol_colors.append('#EA4335')
                            else:
                                vol_colors.append('#4285F4')
                                
                    fig.add_trace(
                        go.Bar(
                            x=df_chart.index,
                            y=df_chart['Volume'],
                            marker_color=vol_colors,
                            name="거래량"
                        ),
                        row=2, col=1
                    )
                    
                    # 20일 거래량 이평선 추가
                    fig.add_trace(
                        go.Scatter(
                            x=df_chart.index,
                            y=df_chart['Vol_SMA_20'],
                            line=dict(color='#8AB4F8', width=1),
                            name="거래량 20MA"
                        ),
                        row=2, col=1
                    )
                    
                    fig.update_layout(
                        height=600,
                        title_text=f"{selected_stock_name} ({selected_ticker}) 주가 & 거래량 분석",
                        hovermode="x unified",
                        xaxis_rangeslider_visible=False,
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # --- 규칙 진단 결과 카드 레이아웃 ---
                    st.markdown("##### 🔍 미너비니의 규칙 자가진단표")
                    col_r1, col_r2 = st.columns(2)
                    
                    def get_status_str(cond):
                        return "🟢 충족 (Pass)" if cond else "🔴 미흡 (Fail)"
                        
                    with col_r1:
                        st.write(f"1. 주가 > 150MA & 200MA: {get_status_str(cond1)}  *(현재가: {fmt_curr(current_price, selected_ticker)} | 150MA: {fmt_curr(ma150_val, selected_ticker)} | 200MA: {fmt_curr(ma200_val, selected_ticker)})*")
                        st.write(f"2. 이평선 우상향 정배열 (150MA > 200MA): {get_status_str(cond2)}  *(150MA: {fmt_curr(ma150_val, selected_ticker)} > 200MA: {fmt_curr(ma200_val, selected_ticker)})*")
                        st.write(f"3. 200MA 한 달 이상 상승세: {get_status_str(cond3)}  *(현재 200MA: {fmt_curr(ma200_val, selected_ticker)} | 한달 전: {fmt_curr(ma200_prev_val, selected_ticker)})*")
                        st.write(f"4. 50MA가 장기 이평 위에 위치: {get_status_str(cond4)}  *(50MA: {fmt_curr(ma50_val, selected_ticker)} | 150MA: {fmt_curr(ma150_val, selected_ticker)})*")
                        
                    with col_r2:
                        st.write(f"5. 현재 주가가 50MA 위에 위치: {get_status_str(cond5)}  *(현재가: {fmt_curr(current_price, selected_ticker)} | 50MA: {fmt_curr(ma50_val, selected_ticker)})*")
                        st.write(f"6. 52주 신저가 대비 30% 이상 상승: {get_status_str(cond6)}  *(현재가: {fmt_curr(current_price, selected_ticker)} | 52주 최저: {fmt_curr(l52_val, selected_ticker)} | 상승폭: +{row['Pct_Above_Low']:.2f}%)*")
                        st.write(f"7. 52주 신고가 대비 25% 이내 근접: {get_status_str(cond7)}  *(현재가: {fmt_curr(current_price, selected_ticker)} | 52주 최고: {fmt_curr(h52_val, selected_ticker)} | 낙폭: -{row['Pct_Below_High']:.2f}%)*")
                        st.write(f"8. 상대강도(RS Rating) 요건 (70 이상): {get_status_str(cond8)}  *(개별 종목 RS Rating: **{stock_rs}** | 기준값: {rs_rating_thresh})*")
                else:
                    st.error("해당 종목의 상세 차트 데이터를 가져오는데 실패했습니다.")
    else:
        st.info("💡 사이드바에서 조건을 확인한 뒤 **'🚀 스크리닝 시작'** 버튼을 클릭하시면 포착된 초성장 종목의 리스트와 차트가 표시됩니다.")

# ----------------- 탭 2: 개별 관심 종목 수동 분석기 -----------------
with tab2:
    st.markdown("### 🔬 개별 관심 종목 분석기 (수동 조회)")
    st.markdown("스크리닝을 거치지 않더라도, 조회하고자 하는 개별 종목 티커를 입력하여 미너비니의 **상승 추세 조건 충족 상황**과 **최근 60일간 변동성(VCP) 수축 모습**을 상세 진단할 수 있습니다.")
    
    # 한국 및 미국 샘플 예시 가이드
    st.info("💡 **티커 입력 예시:**\n- 한국 코스피: `005930.KS` (삼성전자) | 코스닥: `247540.KQ` (에코프로비엠)\n- 미국 주식: `AAPL` (애플), `NVDA` (엔비디아), `TSLA` (테슬라)")
    
    manual_ticker = st.text_input("분석할 종목의 티커를 입력하세요:", value="NVDA").strip().upper()
    
    if manual_ticker:
        with st.spinner(f"관심 종목 {manual_ticker} 분석 중..."):
            # 차트/분석을 위해 2년 데이터 획득
            df_m = yf.download(manual_ticker, period="2y", progress=False)
            if isinstance(df_m.columns, pd.MultiIndex):
                df_m.columns = df_m.columns.droplevel(1)
            df_m = df_m.dropna(subset=['Close', 'High', 'Low', 'Volume'])
            
        if not df_m.empty:
            df_m['SMA_50'] = df_m['Close'].rolling(window=50).mean()
            df_m['SMA_150'] = df_m['Close'].rolling(window=150).mean()
            df_m['SMA_200'] = df_m['Close'].rolling(window=200).mean()
            df_m['Vol_SMA_20'] = df_m['Volume'].rolling(window=20).mean()
            df_m['High_52w'] = df_m['High'].rolling(window=252).max()
            df_m['Low_52w'] = df_m['Low'].rolling(window=252).min()
            
            curr_p = df_m['Close'].iloc[-1]
            ma50_m = df_m['SMA_50'].iloc[-1]
            ma150_m = df_m['SMA_150'].iloc[-1]
            ma200_m = df_m['SMA_200'].iloc[-1]
            ma200_prev_m = df_m['SMA_200'].iloc[-22] if len(df_m) >= 22 else df_m['SMA_200'].iloc[0]
            h52_m = df_m['High_52w'].iloc[-1]
            l52_m = df_m['Low_52w'].iloc[-1]
            
            c_below_high = ((h52_m - curr_p) / h52_m) * 100.0
            c_above_low = ((curr_p - l52_m) / l52_m) * 100.0
            
            # 8대 규칙 점검 (RS Rating은 전체 스크리닝이 구동 완료되어 전체 분포 리스트가 있는 경우에만 정확히 대조)
            cond1_m = (curr_p > ma150_m) and (curr_p > ma200_m)
            cond2_m = ma150_m > ma200_m
            cond3_m = ma200_m > ma200_prev_m
            cond4_m = (ma50_m > ma150_m) and (ma50_m > ma200_m)
            cond5_m = curr_p > ma50_m
            cond6_m = curr_p >= (l52_m * 1.30)
            cond7_m = curr_p >= (h52_m * 0.75)
            
            # RS rating 가져오기 시도
            manual_rs = st.session_state.rs_ratings.get(manual_ticker, None)
            if manual_rs is not None:
                cond8_m = manual_rs >= rs_rating_thresh
                rs_rating_str = f"**{manual_rs}**"
            else:
                cond8_m = None
                rs_rating_str = "조회 필요 *(메인 스크리너 구동 후 판단 가능)*"
                
            # VCP 조건 검토
            amp1_m, is_vcp_m = check_vcp_pattern(df_m, amp_limit=vcp_amp_limit, vol_dryup_ratio=vol_dryup_ratio, breakout_pct=breakout_pct)
            
            # 1. 차트 플로팅
            fig_m = make_subplots(
                rows=2, cols=1,
                shared_xaxes=True,
                vertical_spacing=0.08,
                row_heights=[0.7, 0.3]
            )
            
            fig_m.add_trace(
                go.Candlestick(
                    x=df_m.index, open=df_m['Open'], high=df_m['High'], low=df_m['Low'], close=df_m['Close'],
                    name="주가", increasing_line_color='#EA4335', decreasing_line_color='#4285F4'
                ),
                row=1, col=1
            )
            
            fig_m.add_trace(go.Scatter(x=df_m.index, y=df_m['SMA_50'], line=dict(color='#FBBC05', width=1.5), name="50일 MA"), row=1, col=1)
            fig_m.add_trace(go.Scatter(x=df_m.index, y=df_m['SMA_150'], line=dict(color='#34A853', width=1.5), name="150일 MA"), row=1, col=1)
            fig_m.add_trace(go.Scatter(x=df_m.index, y=df_m['SMA_200'], line=dict(color='#EA4335', width=2), name="200일 MA"), row=1, col=1)
            fig_m.add_trace(go.Scatter(x=df_m.index, y=df_m['High_52w'], line=dict(color='#BDC1C6', width=1, dash='dash'), name="52주 신고가선"), row=1, col=1)
            fig_m.add_trace(go.Scatter(x=df_m.index, y=df_m['Low_52w'], line=dict(color='#BDC1C6', width=1, dash='dot'), name="52주 신저가선"), row=1, col=1)
            
            # VCP 3구간 색상 칠하기
            if len(df_m) >= 60:
                d60_m = df_m.index[-60]
                d30_m = df_m.index[-30]
                d10_m = df_m.index[-10]
                dend_m = df_m.index[-1]
                
                m_p3 = df_m.iloc[-60:-30]
                m_p2 = df_m.iloc[-30:-10]
                m_p1 = df_m.iloc[-10:]
                
                m_amp3 = (m_p3['High'].max() - m_p3['Low'].min()) / m_p3['Low'].min() * 100.0
                m_amp2 = (m_p2['High'].max() - m_p2['Low'].min()) / m_p2['Low'].min() * 100.0
                m_amp1 = (m_p1['High'].max() - m_p1['Low'].min()) / m_p1['Low'].min() * 100.0
                
                fig_m.add_vrect(x0=d60_m, x1=d30_m, fillcolor="rgba(66, 133, 244, 0.05)", line_width=0, annotation_text=f"Amp3: {m_amp3:.1f}%", annotation_position="top left", row=1, col=1)
                fig_m.add_vrect(x0=d30_m, x1=d10_m, fillcolor="rgba(251, 188, 5, 0.05)", line_width=0, annotation_text=f"Amp2: {m_amp2:.1f}%", annotation_position="top left", row=1, col=1)
                fig_m.add_vrect(x0=d10_m, x1=dend_m, fillcolor="rgba(52, 168, 83, 0.08)", line_width=0, annotation_text=f"Amp1: {m_amp1:.1f}%", annotation_position="top left", row=1, col=1)
                
            # 거래량 보조 지표
            v_colors = []
            for idx in range(len(df_m)):
                if idx == 0:
                    v_colors.append('#EA4335')
                else:
                    if df_m['Close'].iloc[idx] >= df_m['Close'].iloc[idx-1]:
                        v_colors.append('#EA4335')
                    else:
                        v_colors.append('#4285F4')
                        
            fig_m.add_trace(go.Bar(x=df_m.index, y=df_m['Volume'], marker_color=v_colors, name="거래량"), row=2, col=1)
            fig_m.add_trace(go.Scatter(x=df_m.index, y=df_m['Vol_SMA_20'], line=dict(color='#8AB4F8', width=1), name="거래량 20MA"), row=2, col=1)
            
            fig_m.update_layout(
                height=600,
                title_text=f"{manual_ticker} 주가 & 거래량 분석 (관심 종목)",
                hovermode="x unified",
                xaxis_rangeslider_visible=False,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig_m, use_container_width=True)
            
            # 2. 결과 종합 진단 표
            st.markdown("##### ⚙️ 종목 상세 상태 검토")
            col_m1, col_m2 = st.columns(2)
            
            def get_status_str(cond):
                if cond is None:
                    return "⚪ 판정 불가 (RS 데이터 누락)"
                return "🟢 충족 (Pass)" if cond else "🔴 미협 (Fail)"
                
            with col_m1:
                st.write(f"1. 주가 > 150MA & 200MA: {get_status_str(cond1_m)}  *(현재가: {fmt_curr(curr_p, manual_ticker)} | 150MA: {fmt_curr(ma150_m, manual_ticker)} | 200MA: {fmt_curr(ma200_m, manual_ticker)})*")
                st.write(f"2. 이평선 우상향 정배열 (150MA > 200MA): {get_status_str(cond2_m)}  *(150MA: {fmt_curr(ma150_m, manual_ticker)} > 200MA: {fmt_curr(ma200_m, manual_ticker)})*")
                st.write(f"3. 200MA 한 달 이상 상승세: {get_status_str(cond3_m)}  *(현재 200MA: {fmt_curr(ma200_m, manual_ticker)} | 한달 전: {fmt_curr(ma200_prev_m, manual_ticker)})*")
                st.write(f"4. 50MA가 장기 이평 위에 위치: {get_status_str(cond4_m)}  *(50MA: {fmt_curr(ma50_m, manual_ticker)} | 150MA: {fmt_curr(ma150_m, manual_ticker)})*")
                
            with col_m2:
                st.write(f"5. 현재 주가가 50MA 위에 위치: {get_status_str(cond5_m)}  *(현재가: {fmt_curr(curr_p, manual_ticker)} | 50MA: {fmt_curr(ma50_m, manual_ticker)})*")
                st.write(f"6. 52주 신저가 대비 30% 이상 상승: {get_status_str(cond6_m)}  *(현재가: {fmt_curr(curr_p, manual_ticker)} | 52주 최저: {fmt_curr(l52_m, manual_ticker)} | 상승폭: +{c_above_low:.2f}%)*")
                st.write(f"7. 52주 신고가 대비 25% 이내 근접: {get_status_str(cond7_m)}  *(현재가: {fmt_curr(curr_p, manual_ticker)} | 52주 최고: {fmt_curr(h52_m, manual_ticker)} | 낙폭: -{c_below_high:.2f}%)*")
                st.write(f"8. 상대강도(RS Rating) 요건 (70 이상): {get_status_str(cond8_m)}  *(개별 종목 RS Rating: **{rs_rating_str}**)*")
                
            st.markdown("---")
            if is_vcp_m:
                st.success(f"🌀 **VCP 조건 종합 판정:** **합격(Pass)**! 최근 10일간 진폭(Amp1)이 {amp1_m*100.0:.2f}%로 좁혀지며 에너지가 수축되었습니다.")
            else:
                st.warning("🌀 **VCP 조건 종합 판정:** **미흡(Fail)**. 3단계 진폭 순차적 수축 요건(Amp3 > Amp2 > Amp1)을 충족하지 못했거나, 거래량 Dry-up 혹은 가격이 돌파 임박 지점(최근 20일 최고가의 설정한 비율 이상)이 아닙니다.")
        else:
            st.error(f"티커 '{manual_ticker}'에 대한 데이터를 불러올 수 없습니다. 티커명을 다시 확인해 주세요.")
